import csv
import openpyxl
from openpyxl.styles import Font
from show_logic import find_show

def export_cues_to_csv(show_id: int, file_path: str):
    """
    Exportiert die Cuelist für den 'Generic CSV' Import von EOS.
    Format: 'Cue', 'Label', 'Notes', 'Up Time', 'Down Time'
    Wichtig: 'Cue' Spalte enthält 'ListenNummer/CueNummer' (z.B. 1/10).
    Import in EOS via: File -> Import -> CSV -> Cues (nicht Console Data!)
    """
    show = find_show(show_id)
    if not show:
        raise ValueError("Show not found")
        
    songs = show.get("songs", [])
    cuelist_id = show.get("eos_cuelist_id", 1)
    
    # Simple Header für Generic CSV Import
    headers = ["Cue", "Label", "Notes", "Up Time", "Down Time"]
    
    # Windows-Style Zeilenenden (\r\n) und UTF-8 BOM
    with open(file_path, "w", newline='\r\n', encoding='utf-8-sig') as csvfile:
        writer = csv.writer(csvfile, delimiter=',')
        writer.writerow(headers)
        
        for idx, song in enumerate(songs, 1):
            cue_num = song.get("order_index", idx)
            name = song.get("name", f"Cue {cue_num}")
            mood = song.get("mood", "")
            colors = song.get("colors", "")
            notes = (song.get("special_notes") or "") + " " + (song.get("general_notes") or "")
            
            # WICHTIG: Listen-Nummer/Cue-Nummer Format (z.B. "1/10")
            cue_ident = f"{cuelist_id}/{cue_num}"
            
            label = name
            if mood or colors:
                label += f" [{mood}|{colors}]"
                
            writer.writerow([
                cue_ident,
                label.strip(),
                notes.strip(),
                "", # Up Time default
                ""  # Down Time default
            ])

def export_cues_to_xlsx(show_id: int, file_path: str):
    """
    Exportiert die Cuelist als Excel-Datei im 'Simple' Format.
    Import in EOS via: File -> Import -> CSV -> Cues (XLSX wählen).
    """
    show = find_show(show_id)
    if not show:
        raise ValueError("Show not found")
        
    songs = show.get("songs", [])
    cuelist_id = show.get("eos_cuelist_id", 1)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EOS_Cues"
    
    # Header
    headers = ["Cue", "Label", "Notes", "Up Time", "Down Time"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        
    # Daten
    for idx, song in enumerate(songs, 2):
        cue_num = song.get("order_index", idx - 1)
        name = song.get("name", f"Cue {cue_num}")
        mood = song.get("mood", "")
        colors = song.get("colors", "")
        notes = (song.get("special_notes") or "") + " " + (song.get("general_notes") or "")
        
        # WICHTIG: Listen-Nummer/Cue-Nummer Format
        cue_ident = f"{cuelist_id}/{cue_num}"
        
        label = name
        if mood or colors:
            label += f" [{mood}|{colors}]"
            
        ws.cell(row=idx, column=1, value=cue_ident)
        ws.cell(row=idx, column=2, value=label.strip())
        ws.cell(row=idx, column=3, value=notes.strip())
        ws.cell(row=idx, column=4, value="") # Up Time
        ws.cell(row=idx, column=5, value="") # Down Time
        
    wb.save(file_path)
